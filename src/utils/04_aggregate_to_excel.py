
# 04_aggregate_to_excel.py
# ========================
# Step 4 of the Fondazione del Monte extraction pipeline.
 
# Reads all .json files from a folder and produces a single Excel workbook
# with four sheets:
 
#   Sheet 1 — "Dati per progetto"
#     One row per project, one column per indicator (values only).
#     This is your main working sheet.
 
#   Sheet 2 — "Aggregati"
#     Summary statistics across all projects:
#     - Numeric fields: summed across all projects
#     - Category fields (area, durata, frequenza, etc.): count per category
#     - SI/NO fields (leve, missioni): count of SI across projects
#     - Organisation lists: unique organisations counted and ranked
 
#   Sheet 3 — "Confidenza"
#     Same layout as Sheet 1 but showing the confidence score per field,
#     colour-coded: GREEN=HIGH, ORANGE=MEDIUM, RED=LOW, GREY=null.
 
#   Sheet 4 — "Fonti"
#     Same layout as Sheet 1 but showing the source snippets for verification.
 
# HOW TO USE
# ----------
# 1. Save all your .json files into a folder called "json_outputs"
#    (in the same directory as this script).
#    Name them anything you like — e.g. casa_adolescente.json, glab.json etc.
 
# 2. Run:
#        python 04_aggregate_to_excel.py
 
# 3. Output: aggregated_results.xlsx

 
import json
import os
import sys
from pathlib import Path
from collections import Counter
 
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# ── Configuration ─────────────────────────────────────────────────────────────
 
JSON_DIR    = "src/utils/json_outputs"
OUTPUT_FILE = "src/utils/aggregated_results.xlsx"
 
# ── Field definitions ─────────────────────────────────────────────────────────
 
# Field key → (short label for column header, aggregation type)
# Types: "text", "numeric", "category", "list", "sino"
FIELDS = {
    "project_name":          ("Nome progetto",           "text"),
    "organisation_name":     ("Organizzazione",          "text"),
    "01_aree_tematiche":     ("01 Aree tematiche",       "category"),
    "02_programma_raggruppamento": ("02 Programma/Bando","category"),
    "03_ruolo_fondazione":   ("03 Ruolo Fondazione",     "category"),
    "04_budget_complessivo": ("04 Budget totale (€)",    "numeric"),
    "05_quota_fondazione":   ("05 Quota Fondazione (€)", "numeric"),
    "06_risorse_non_monetarie_fondazione": ("06 Risorse non monetarie", "text"),
    "07_quota_altre_org":    ("07 Quota altre org (€)",  "numeric"),
    "08_org_cofinanziatrici":("08 Org. co-finanziatrici","list"),
    "09_org_partner_non_fin":("09 Org. partner non fin.","list"),
    "10_durata":             ("10 Durata",               "category"),
    "11_frequenza":          ("11 Frequenza",            "category"),
    "12_livello_territoriale":("12 Livello territoriale","category"),
    "13_n_coop_sociali":     ("13 N. coop. sociali",     "numeric"),
    "14_n_associazioni_odv": ("14 N. assoc. e OdV",      "numeric"),
    "15_n_altre_org_nonprofit":("15 N. altre org NP",    "numeric"),
    "16_n_ist_pubbliche_territorio":("16 N. ist. pubbl. territorio","numeric"),
    "17_n_scuole_ist_formazione":("17 N. scuole/ist. form.","numeric"),
    "18_n_universita":       ("18 N. università",        "numeric"),
    "19_n_altre_ist_pubbliche":("19 N. altre ist. pubbl.","numeric"),
    "20_n_altri_soggetti":   ("20 N. altri soggetti",    "numeric"),
    "21_n_persone_vulnerabilita_economica": ("21 Vuln. economica",  "numeric"),
    "22_n_persone_disabilita":("22 Disabilità",           "numeric"),
    "23_n_persone_vulnerabilita_psicosociale":("23 Vuln. psicosociale","numeric"),
    "24_n_persone_vulnerabilita_migratoria":("24 Vuln. migratoria", "numeric"),
    "25_n_persone_altre_vulnerabilita":("25 Altre vuln.",  "numeric"),
    "26_n_persone_vulnerabilita_multipla":("26 Vuln. multipla","numeric"),
    "27_n_minori_0_18_vulnerabili":("27 Minori 0-18 vuln.","numeric"),
    "28_n_giovani_adulti_18_35_vulnerabili":("28 Giovani 18-35 vuln.","numeric"),
    "29_n_adulti_36_65_vulnerabili":("29 Adulti 36-65 vuln.","numeric"),
    "30_n_anziani_65plus_vulnerabili":("30 Anziani >65 vuln.","numeric"),
    "31_n_donne_vulnerabili":("31 Donne vuln.",           "numeric"),
    "32_n_persone_non_vulnerabili":("32 Non vulnerabili", "numeric"),
    "33_n_minori_0_18_non_vulnerabili":("33 Minori 0-18 non vuln.","numeric"),
    "34_n_giovani_adulti_18_35_non_vulnerabili":("34 Giovani 18-35 non vuln.","numeric"),
    "35_n_adulti_36_65_non_vulnerabili":("35 Adulti 36-65 non vuln.","numeric"),
    "36_n_anziani_65plus_non_vulnerabili":("36 Anziani >65 non vuln.","numeric"),
    "37_n_donne_non_vulnerabili":("37 Donne non vuln.",   "numeric"),
    "38_leva_1_scienza_tecnologia_cultura":("38 Leva 1: Scienza/tecnologia/cultura","sino"),
    "39_leva_2_capacity_building":("39 Leva 2: Capacity Building","sino"),
    "40_leva_3_azione_individuale_collettiva":("40 Leva 3: Azione ind./collettiva","sino"),
    "41_leva_4_governance":  ("41 Leva 4: Governance",   "sino"),
    "42_leva_5_finanza_economia":("42 Leva 5: Finanza/economia","sino"),
    "43_missione_1a_beni_ambientali":("43 Miss. 1a: Beni ambientali","sino"),
    "44_missione_1b_bisogni_base":("44 Miss. 1b: Bisogni base","sino"),
    "45_missione_2a_gestione_risorse":("45 Miss. 2a: Gestione risorse","sino"),
    "46_missione_2b_contrasto_disuguaglianze":("46 Miss. 2b: Contrasto disuguaglianze","sino"),
    "47_n_org_beneficiarie_totale":("47 Tot. org. beneficiarie","numeric"),
    "48_org_beneficiarie_per_leva_missione":("48 Org. per leva/missione","text"),
    "49_n_persone_beneficiarie_totale":("49 Tot. persone beneficiarie","numeric"),
    "50_persone_beneficiarie_per_leva_missione":("50 Persone per leva/missione","text"),
}
 
# Aliases: some JSON files may use slightly different key names
KEY_ALIASES = {
    "07_quota_altre_organizzazioni": "07_quota_altre_org",
    "08_org_cofinanziatrici":        "08_org_cofinanziatrici",
    "09_org_partner_non_finanziatrici": "09_org_partner_non_fin",
}
 
# ── Styles ────────────────────────────────────────────────────────────────────
 
HEADER_FILL   = PatternFill("solid", start_color="2F4F8F")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBHEAD_FILL  = PatternFill("solid", start_color="D9E1F2")
SUBHEAD_FONT  = Font(bold=True, name="Arial", size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN  = Alignment(horizontal="center", vertical="top")
THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
 
CONF_COLORS = {
    "HIGH":   "C6EFCE",   # green
    "MEDIUM": "FFEB9C",   # orange/yellow
    "LOW":    "FFC7CE",   # red
    None:     "F2F2F2",   # grey
}
 
# ── Helpers ───────────────────────────────────────────────────────────────────
 
def load_json_files(json_dir: Path) -> list[dict]:
    projects = []
    for f in sorted(json_dir.glob("*.json")):
        try:
            text = f.read_text(encoding="utf-8")
            data = json.loads(text)
            data["_filename"] = f.name
            projects.append(data)
            print(f"  ✓  Loaded: {f.name}")
        except json.JSONDecodeError as e:
            print(f"  ✗  Invalid JSON in {f.name}: {e}")
        except Exception as e:
            print(f"  ✗  Could not read {f.name}: {e}")
    return projects
 
 
def resolve_key(data: dict, key: str):
    """Get a field from data, trying the key and any known aliases."""
    if key in data:
        return data[key]
    # Try reverse alias lookup
    for original, alias in KEY_ALIASES.items():
        if alias == key and original in data:
            return data[original]
    return None
 
 
def get_value(data: dict, key: str):
    """Extract the value from a field (handles flat strings and nested dicts)."""
    if key in ("project_name", "organisation_name"):
        return data.get(key)
    field = resolve_key(data, key)
    if field is None:
        return None
    if isinstance(field, dict):
        return field.get("value")
    return field
 
 
def get_confidence(data: dict, key: str) -> str | None:
    field = resolve_key(data, key)
    if isinstance(field, dict):
        return field.get("confidence")
    return None
 
 
def get_source(data: dict, key: str) -> str | None:
    field = resolve_key(data, key)
    if isinstance(field, dict):
        return field.get("source")
    return None
 
 
def style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
 
 
def style_data_cell(cell, numeric=False):
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.alignment = CENTER_ALIGN if numeric else WRAP_ALIGN
 
 
def auto_width(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_w), max_w)
 
 
# ── Sheet 1: Per-project data ─────────────────────────────────────────────────
 
def build_sheet_data(wb, projects):
    ws = wb.active
    ws.title = "Dati per progetto"
 
    keys = list(FIELDS.keys())
    headers = [FIELDS[k][0] for k in keys]
 
    # Header row
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "C2"
 
    for proj in projects:
        row = []
        for key in keys:
            val = get_value(proj, key)
            row.append(val if val is not None else "")
        ws.append(row)
 
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, cell in enumerate(row):
            ftype = FIELDS[keys[i]][1]
            style_data_cell(cell, numeric=(ftype == "numeric"))
 
    auto_width(ws)
    ws.row_dimensions[1].height = 45
 
 
# ── Sheet 2: Aggregates ───────────────────────────────────────────────────────
 
def build_sheet_aggregates(wb, projects):
    ws = wb.create_sheet("Aggregati")
 
    def section_header(title):
        ws.append([title])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        ws.merge_cells(
            start_row=ws.max_row, start_column=1,
            end_row=ws.max_row, end_column=3
        )
        ws.append([])
 
    def sub_header(labels):
        ws.append(labels)
        for col in range(1, len(labels) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = SUBHEAD_FILL
            cell.font = SUBHEAD_FONT
            cell.border = THIN_BORDER
 
    def data_row(values):
        ws.append(values)
        for col in range(1, len(values) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
 
    n = len(projects)
    ws.append([f"Aggregati su {n} progetti"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, name="Arial", size=12)
    ws.append([])
 
    # ── 1. Numeric sums ──
    section_header("TOTALI NUMERICI")
    sub_header(["Indicatore", "Totale (su tutti i progetti)", "N. progetti con dato"])
    numeric_keys = [k for k, (_, t) in FIELDS.items() if t == "numeric"]
    for key in numeric_keys:
        label = FIELDS[key][0]
        values = [get_value(p, key) for p in projects]
        nums = [v for v in values if v is not None and str(v).strip() not in ("", "null")]
        try:
            total = sum(float(v) for v in nums)
            total_str = int(total) if total == int(total) else round(total, 2)
        except Exception:
            total_str = "N/D"
        data_row([label, total_str, f"{len(nums)}/{n}"])
 
    ws.append([])
 
    # ── 2. Category breakdowns ──
    section_header("DISTRIBUZIONE PER CATEGORIA")
    category_keys = [k for k, (_, t) in FIELDS.items() if t == "category"]
    for key in category_keys:
        label = FIELDS[key][0]
        sub_header([label, "N. progetti", "% progetti"])
        values = [get_value(p, key) for p in projects]
        # Each value may contain multiple categories separated by |
        all_cats = []
        for v in values:
            if v:
                for cat in str(v).split("|"):
                    cat = cat.strip()
                    if cat:
                        all_cats.append(cat)
        counter = Counter(all_cats)
        for cat, count in sorted(counter.items(), key=lambda x: -x[1]):
            pct = f"{round(count / n * 100)}%"
            data_row([f"  {cat}", count, pct])
        ws.append([])
 
    # ── 3. SI/NO fields ──
    section_header("LEVE E MISSIONI (N. PROGETTI CON SI)")
    sub_header(["Leva / Missione", "N. progetti SI", "% progetti SI"])
    sino_keys = [k for k, (_, t) in FIELDS.items() if t == "sino"]
    for key in sino_keys:
        label = FIELDS[key][0]
        values = [get_value(p, key) for p in projects]
        si_count = sum(1 for v in values if v and str(v).strip().upper() == "SI")
        pct = f"{round(si_count / n * 100)}%"
        data_row([label, si_count, pct])
 
    ws.append([])
 
    # ── 4. Organisation lists ──
    section_header("ORGANIZZAZIONI PIÙ FREQUENTI")
    for key in ("08_org_cofinanziatrici", "09_org_partner_non_fin"):
        label = FIELDS.get(key, (key,))[0]
        sub_header([label, "N. progetti in cui appare", ""])
        all_orgs = []
        for p in projects:
            val = get_value(p, key)
            if val:
                for org in str(val).split("|"):
                    org = org.strip()
                    if org:
                        all_orgs.append(org)
        counter = Counter(all_orgs)
        for org, count in sorted(counter.items(), key=lambda x: -x[1]):
            data_row([f"  {org}", count, ""])
        ws.append([])
 
    # Column widths
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15
 
 
# ── Sheet 3: Confidence ───────────────────────────────────────────────────────
 
def build_sheet_confidence(wb, projects):
    ws = wb.create_sheet("Confidenza")
    keys = list(FIELDS.keys())
    headers = [FIELDS[k][0] for k in keys]
 
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "C2"
 
    for proj in projects:
        row_vals = []
        for key in keys:
            if key in ("project_name", "organisation_name"):
                row_vals.append(get_value(proj, key) or "")
            else:
                conf = get_confidence(proj, key)
                row_vals.append(conf or "")
        ws.append(row_vals)
 
    # Colour-code confidence cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, cell in enumerate(row):
            key = keys[i]
            if key not in ("project_name", "organisation_name"):
                conf = cell.value
                color = CONF_COLORS.get(conf, CONF_COLORS[None])
                cell.fill = PatternFill("solid", start_color=color)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
 
    auto_width(ws)
    ws.row_dimensions[1].height = 45
 
    # Legend
    ws.append([])
    ws.append(["LEGENDA COLORI:"])
    legend_row = ws.max_row
    ws.cell(legend_row, 1).font = Font(bold=True, name="Arial")
    for i, (conf, color) in enumerate(CONF_COLORS.items(), 1):
        label = conf if conf else "non disponibile"
        ws.append([f"  {label}"])
        cell = ws.cell(ws.max_row, 1)
        cell.fill = PatternFill("solid", start_color=color)
        cell.font = NORMAL_FONT
 
 
# ── Sheet 4: Sources ──────────────────────────────────────────────────────────
 
def build_sheet_sources(wb, projects):
    ws = wb.create_sheet("Fonti")
    keys = list(FIELDS.keys())
    headers = [FIELDS[k][0] for k in keys]
 
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "C2"
 
    for proj in projects:
        row_vals = []
        for key in keys:
            if key in ("project_name", "organisation_name"):
                row_vals.append(get_value(proj, key) or "")
            else:
                src = get_source(proj, key)
                row_vals.append(src or "")
        ws.append(row_vals)
 
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
 
    auto_width(ws)
    ws.row_dimensions[1].height = 45
 
 
# ── Main ──────────────────────────────────────────────────────────────────────
 
def main():
    json_dir = Path(JSON_DIR)
 
    if not json_dir.exists():
        print(f"\nERROR: Folder '{JSON_DIR}' not found.")
        print("Create it and place all your .json files inside it.")
        sys.exit(1)
 
    print(f"\nLoading JSON files from '{JSON_DIR}'...")
    projects = load_json_files(json_dir)
 
    if not projects:
        print("No valid JSON files found. Exiting.")
        sys.exit(1)
 
    print(f"\nBuilding Excel workbook for {len(projects)} project(s)...")
 
    wb = Workbook()
    build_sheet_data(wb, projects)
    build_sheet_aggregates(wb, projects)
    build_sheet_confidence(wb, projects)
    build_sheet_sources(wb, projects)
 
    out_path = Path(OUTPUT_FILE)
    wb.save(str(out_path))
    size_kb = out_path.stat().st_size / 1024
    print(f"\n✓  Saved → {out_path}  ({size_kb:.1f} KB)")
    print("\nSheets created:")
    print("  1. Dati per progetto  — one row per project, all values")
    print("  2. Aggregati          — sums, category breakdowns, org counts")
    print("  3. Confidenza         — colour-coded confidence scores")
    print("  4. Fonti              — source snippets for verification")
 
 
if __name__ == "__main__":
    main()