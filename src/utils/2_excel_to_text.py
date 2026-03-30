# 02_excel_to_text.py
# ===================
# Step 2 of the Fondazione del Monte extraction pipeline.
 
# Reads all Excel files (.xlsx, .xls) in each project folder and appends
# their content to the project's existing .txt file (created by script 01).
 
# HOW TO USE
# ----------
# 1. Install dependencies:
#         pip install openpyxl
#         (only needed for legacy .xls files: pip install xlrd)
 
# 2. Make sure script 01 has already been run first.
 
# 3. Run:
#         python 02_excel_to_text.py
 
# RE-RUNNING SAFELY
# -----------------
# If you need to re-run (e.g. you added new files to a project folder):
 
#   - To reprocess ALL projects from scratch:
#         Set OVERWRITE_MODE = True  (deletes and recreates all .txt files)
#         Then re-run script 01 first, then script 02.
 
#   - To reprocess just ONE project:
#         Set SINGLE_PROJECT = "nome_cartella_progetto"
#         This will delete and recreate only that project's .txt file.
#         Then re-run script 01 with the same SINGLE_PROJECT setting,
#         then re-run script 02.
 
# FILE TYPES HANDLED
# ------------------
# Two Excel formats are present across the projects:
 
#   1. "Scheda aggiuntiva" (Q&A format):
#      - Sheet "Modulo di rendicontazione"
#      - Col B = question, Col C = answer, Col D = evidence, Col E = extra data
#      - All filled columns are extracted.
#      - Sheet "Liste" is skipped (dropdown options only).
 
#   2. "Parte 2" financial accountability sheets:
#      - Contain payslips, invoices, budget totals, income sources.
#      - Read as a flat table — category headers + line items + totals.
#      - Useful for indicators like total budget, staff costs, funding sources.
#      - Detail sheets (individual payslip breakdowns) are included but
#        clearly labelled so the LLM can distinguish summary from detail.

 
import sys
from pathlib import Path
import openpyxl
 
# ── Configuration ─────────────────────────────────────────────────────────────
 
INPUT_DIR  = "src/utils/projects"
OUTPUT_DIR = "src/utils/output_texts"
 
# Set to True to delete and recreate ALL project .txt files on next run.
# Use this if you've added/changed files across multiple projects.
OVERWRITE_MODE = False
 
# Set to a project folder name (e.g. "casa_adolescente") to reprocess
# only that one project. Set to None to process all projects.
# When set, this project's .txt file is deleted and recreated from scratch.
SINGLE_PROJECT = None   # e.g. "casa_adolescente"
 
# Sheet names to always skip — these are dropdown/reference lists only
SKIP_SHEETS = {"liste", "liste a tendina", "dropdown", "reference",
               "riferimenti", "legenda"}
 
# Column indices (0-based): A=0, B=1, C=2, D=3, E=4
COL_QUESTION  = 1   # B
COL_ANSWER    = 2   # C
COL_EVIDENCE  = 3   # D
COL_EXTRA     = 4   # E  ← extra data column present in some scheda files
 
# Placeholder texts in evidence/extra cells to skip (template boilerplate)
EVIDENCE_PLACEHOLDERS = {
    "quale dato/evidenza a supporto di questa risposta? se presenti, indicatelo nella cella a dx.",
    "eventuali dati/evidenze a supporto (quando richiesto nella cella)",
    "eeventuali dati/evidenze a supporto (quando richiesto)",
    "eventuali dati/evidenze a supporto (quando richiesto)",
}
 
# ── Helpers ───────────────────────────────────────────────────────────────────
 
def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\n", " ").replace("\r", " ")
 
def is_placeholder(text: str) -> bool:
    return text.strip().lower() in EVIDENCE_PLACEHOLDERS
 
def is_skip_sheet(name: str) -> bool:
    return name.strip().lower() in SKIP_SHEETS
 
def is_qa_sheet(ws) -> bool:
    
    # Detect if a sheet uses the Q&A layout (col B = questions, col C = answers).
    # Checks the first 15 non-empty rows.
    
    b_filled = 0
    checked = 0
    for row in ws.iter_rows(max_row=20, values_only=True):
        row = list(row) + [None] * 6
        if any(v is not None for v in row):
            if clean(row[COL_QUESTION]):
                b_filled += 1
            checked += 1
        if checked >= 15:
            break
    return b_filled >= 3
 
 
def read_qa_sheet(ws) -> str:
    
    # Read a Q&A sheet: col B = question, col C = answer,
    # col D = evidence, col E = extra data.
    # Section headers (question with no answer) are formatted as [Header].
    
    lines = []
    for row in ws.iter_rows(values_only=True):
        row = list(row) + [None] * 6
        question  = clean(row[COL_QUESTION])
        answer    = clean(row[COL_ANSWER])
        evidence  = clean(row[COL_EVIDENCE])
        extra     = clean(row[COL_EXTRA])
 
        if not question and not answer:
            continue
 
        # Section header — question present, no answer
        if question and not answer:
            lines.append(f"\n[{question}]")
            continue
 
        # Q&A row
        if question and answer:
            lines.append(f"Q: {question}")
            lines.append(f"A: {answer}")
        elif answer:
            lines.append(f"A: {answer}")
 
        # Append evidence if meaningful
        if evidence and not is_placeholder(evidence):
            lines.append(f"Evidence: {evidence}")
 
        # Append extra column data if meaningful
        if extra and not is_placeholder(extra):
            lines.append(f"Data: {extra}")
 
        lines.append("")  # blank line between entries
 
    return "\n".join(lines)
 
 
def read_financial_sheet(ws, sheet_name: str) -> str:
    
    # Read a financial accountability sheet as a flat table.
    # Extracts: category headers, line item descriptions, and numeric values.
    # Skips rows that are entirely empty or contain only formula strings.
    
    lines = [f"[Financial sheet: {sheet_name}]"]
    for row in ws.iter_rows(values_only=True):
        cells = [clean(c) for c in row if clean(c)]
        # Skip rows with only formula-like content
        if not cells:
            continue
        if all(c.startswith("=") for c in cells):
            continue
        lines.append("  " + " | ".join(cells))
    return "\n".join(lines)
 
 
def process_excel_file(file_path: Path) -> str:
    
    # Read all relevant sheets from an Excel file and return combined text.
    # Auto-detects Q&A vs financial layout per sheet.
    
    sections = []
 
    try:
        if file_path.suffix.lower() == ".xls":
            try:
                import pandas as pd
                xl = pd.ExcelFile(str(file_path), engine="xlrd")
                for sheet_name in xl.sheet_names:
                    if is_skip_sheet(sheet_name):
                        print(f"    ↷  Skipping '{sheet_name}' (reference list)")
                        continue
                    df = xl.parse(sheet_name, header=None)
                    lines = []
                    for _, row in df.iterrows():
                        cells = [clean(v) for v in row if clean(v)]
                        if cells:
                            lines.append(" | ".join(cells))
                    text = "\n".join(lines)
                    if text.strip():
                        sections.append(f"\n-- Sheet: {sheet_name} --\n{text}")
                return "\n".join(sections)
            except ImportError:
                print(f"    ⚠  xlrd not installed — cannot read {file_path.name}")
                print("       Install with: pip install xlrd")
                return ""
            except Exception as e:
                print(f"    ✗  Failed to read .xls {file_path.name}: {e}")
                return ""
 
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
 
        for sheet_name in wb.sheetnames:
            if is_skip_sheet(sheet_name):
                print(f"    ↷  Skipping '{sheet_name}' (reference list)")
                continue
 
            ws = wb[sheet_name]
            print(f"    → Reading sheet '{sheet_name}' ...")
 
            if is_qa_sheet(ws):
                text = read_qa_sheet(ws)
            else:
                text = read_financial_sheet(ws, sheet_name)
 
            if text.strip():
                sections.append(f"\n-- Sheet: {sheet_name} --\n{text}")
 
        wb.close()
 
    except Exception as e:
        print(f"    ✗  Failed to open {file_path.name}: {e}")
 
    return "\n".join(sections)
 
 
def process_project(project_dir: Path, output_dir: Path, force_overwrite: bool = False):
    
    # Find all Excel files in a project folder and append their content
    # to the project's .txt file.
    # If force_overwrite is True, the existing .txt file is deleted first
    # (you should re-run script 01 before this to rebuild the PDF section).
    
    excel_files = sorted(
        [f for f in project_dir.iterdir()
         if f.suffix.lower() in (".xlsx", ".xls", ".xlsm")]
    )
 
    if not excel_files:
        print(f"  ⚠  No Excel files found in '{project_dir.name}', skipping.")
        return
 
    out_path = output_dir / f"{project_dir.name}.txt"
 
    if force_overwrite and out_path.exists():
        out_path.unlink()
        print(f"  ♻  Deleted existing file for fresh rebuild: {out_path.name}")
 
    all_excel_text = []
    for excel_path in excel_files:
        print(f"  Processing: {excel_path.name}")
        text = process_excel_file(excel_path)
        if text.strip():
            all_excel_text.append(
                f"\n{'─' * 60}\n"
                f"SOURCE FILE: {excel_path.name}\n"
                f"{'─' * 60}\n"
                f"{text}\n"
            )
        else:
            print(f"    ⚠  No content extracted from {excel_path.name}")
 
    if not all_excel_text:
        return
 
    excel_block = (
        f"\n\n{'=' * 60}\n"
        f"EXCEL FILES\n"
        f"{'=' * 60}\n"
        + "\n".join(all_excel_text)
    )
 
    if out_path.exists():
        with open(out_path, "a", encoding="utf-8") as f:
            f.write(excel_block)
        print(f"\n  ✓  Appended Excel content → {out_path}")
    else:
        header = f"PROJECT: {project_dir.name}\n{'=' * 60}\n"
        out_path.write_text(header + excel_block, encoding="utf-8")
        print(f"\n  ✓  Created (no PDF text found) → {out_path}")
 
    size_kb = out_path.stat().st_size / 1024
    print(f"     Total file size: {size_kb:.1f} KB")
 
 
# ── Main ──────────────────────────────────────────────────────────────────────
 
def main():
    input_dir  = Path(INPUT_DIR)
    output_dir = Path(OUTPUT_DIR)
 
    if not input_dir.exists():
        print(f"\nERROR: Input directory '{INPUT_DIR}' not found.")
        sys.exit(1)
 
    output_dir.mkdir(parents=True, exist_ok=True)
 
    # Determine which projects to process
    if SINGLE_PROJECT:
        project_dirs = [input_dir / SINGLE_PROJECT]
        if not project_dirs[0].exists():
            print(f"\nERROR: Project folder '{SINGLE_PROJECT}' not found in '{INPUT_DIR}'.")
            sys.exit(1)
        print(f"\nProcessing single project: {SINGLE_PROJECT}")
    else:
        project_dirs = sorted([d for d in input_dir.iterdir() if d.is_dir()])
        if not project_dirs:
            print(f"No project sub-folders found in '{INPUT_DIR}'.")
            sys.exit(1)
        print(f"\nFound {len(project_dirs)} project(s) in '{INPUT_DIR}'")
 
    if OVERWRITE_MODE:
        print("⚠  OVERWRITE_MODE is ON — existing .txt files will be deleted and rebuilt.")
        print("   Make sure you re-run script 01 before this, or PDF content will be lost.\n")
 
    for project_dir in project_dirs:
        print(f"\n{'━' * 60}")
        print(f"PROJECT: {project_dir.name}")
        print(f"{'━' * 60}")
        process_project(
            project_dir,
            output_dir,
            force_overwrite=(OVERWRITE_MODE or bool(SINGLE_PROJECT))
        )
 
    print(f"\n{'━' * 60}")
    print(f"Done. Text files are in: {output_dir}/")
    print("Next step: upload the .txt files to ChatGPT for indicator extraction.")
 
 
if __name__ == "__main__":
    main()